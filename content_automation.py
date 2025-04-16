
from dotenv import load_dotenv
load_dotenv()

import time
import requests
import pandas as pd
from openai import OpenAI
import logging
import os
import re
import schedule
from datetime import datetime
from cloudinary_utils import upload_image
from instagram_poster import publish_single_post
from instagram_analytics import collect_analytics


# Set up logging
logging.basicConfig(filename='automation.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)

# API Keys (Store these in a .env file)
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
CLOUDINARY_CLOUD_NAME = os.getenv("CLOUDINARY_CLOUD_NAME")
CLOUDINARY_API_KEY = os.getenv("CLOUDINARY_API_KEY")
CLOUDINARY_API_SECRET = os.getenv("CLOUDINARY_API_SECRET")

# File paths
EXCEL_FILE_PATH = "prompts V2.xlsx"
IMAGE_SAVE_PATH = "Generated Images"


# Ensure the image save directory exists
os.makedirs(IMAGE_SAVE_PATH, exist_ok=True)

# Company name for caption generation
COMPANY_NAME = "The Tech Boss"

# Function to sanitize file names
def sanitize_filename(filename):
    filename = re.sub(r'[<>:"/\\|?*]', '', filename)  # Remove invalid characters
    filename = filename.replace(' ', '_')  # Replace spaces with underscores
    filename = filename[:50]  # Limit the length of the filename
    return filename

class CaptionGenerator:
    def __init__(self, openai_api_key):
        self.openai_api_key = openai_api_key
        self.url = "https://api.openai.com/v1/chat/completions"
        self.headers = {
            "Authorization": f"Bearer {self.openai_api_key}",
            "Content-Type": "application/json"
        }

    def generate_caption(self, prompt):
        data = {
            "model": "gpt-4",
            "messages": [
                {"role": "system", "content": prompt},
                {"role": "user", "content": "Please generate a detailed caption and relevant hashtags, ensuring hashtags are included."}
            ]
        }

        for attempt in range(3):  # Retry up to 3 times
            try:
                response = requests.post(self.url, headers=self.headers, json=data)
                if response.status_code == 429:
                    logging.warning(f"Rate limit exceeded. Retrying in {2 ** attempt} seconds...")
                    time.sleep(2 ** attempt)  # Exponential backoff
                    continue  # Retry again
                response.raise_for_status()
                result = response.json()
                content = result.get('choices', [{}])[0].get('message', {}).get('content', "")
                
                if "Hashtags:" in content:
                    caption, hashtags_str = content.split("Hashtags:", 1)
                    caption = caption.strip().replace("\n", " ").replace('"', '')  # Remove quotes
                    hashtags_list = [tag.strip() for tag in hashtags_str.strip().split() if tag.startswith("#")]
                    hashtags = " ".join(hashtags_list)
                    return f"{caption} {hashtags}"
                
                return content.strip().replace("\n", " ").replace('"', '')  # Remove quotes
            
            except requests.exceptions.RequestException as err:
                logging.error(f"Request error: {err}")
                return {"error": str(err)}

        return {"error": "API request failed after multiple retries"}

    def generate_captions_from_excel(self, file_path, company_name):
        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            logging.error(f"Failed to read Excel file: {str(e)}")
            return {"error": f"Failed to read Excel file: {str(e)}"}

        if "prompt" not in df.columns:
            logging.error("Excel file must contain a 'prompt' column.")
            return {"error": "Excel file must contain a 'prompt' column."}

        if "Generated Captions" not in df.columns:
            df["Generated Captions"] = ""  # Initialize the column if it doesn't exist

        new_captions_generated = False
        for index, row in df.iterrows():
            if pd.isna(row["Generated Captions"]) or row["Generated Captions"] == "":
                content = row["prompt"]
                prompt = f"Generate a detailed caption based on {content} and {company_name}, include power words and realism and include 10 relevant hashtags at the end of the caption."
                result = self.generate_caption(prompt)
                
                if isinstance(result, dict) and "error" in result:
                    logging.error(f"Error at row {index}: {result['error']}")
                    df.at[index, "Generated Captions"] = "API Error: Rate limit exceeded"
                else:
                    df.at[index, "Generated Captions"] = result
                    new_captions_generated = True

        if not new_captions_generated:
            logging.info("No new prompts found. Skipping caption generation.")
        else:
            df.to_excel(file_path, index=False)
            logging.info("Captions appended and file saved successfully.")
        return {"status": "Caption generation completed"}

def generate_image(prompt, save_path):
    try:
        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.images.generate(
            model="dall-e-3",
            prompt=prompt,
            size="1024x1024",
            quality="hd",
            n=1,
        )

        image_url = response.data[0].url
        image_data = requests.get(image_url).content
        
        # Sanitize the file name
        file_name = sanitize_filename(prompt[:50])  # Use the first 50 characters of the prompt
        file_path = f"{save_path}/{file_name}.png"
        
        with open(file_path, "wb") as file:
            file.write(image_data)
        logging.info(f"Image saved: {file_path}")
    except Exception as e:
        logging.error(f"Error generating image: {e}")

def generate_images_from_excel(file_path, save_path):
    try:
        df = pd.read_excel(file_path)

        if "Generated Captions" not in df.columns:
            logging.error("The Excel file must contain a 'Generated Captions' column.")
            raise ValueError("The Excel file must contain a 'Generated Captions' column.")

        if "Image URL" not in df.columns:
            df["Image URL"] = ""  # Add a new column for image URLs if it doesn't exist

        new_images_generated = False
        for index, row in df.iterrows():
            prompt = row["Generated Captions"]
            if pd.isna(prompt) or prompt == "":
                continue  # Skip empty captions

            # Check if URL already exists
            if pd.notna(row["Image URL"]) and row["Image URL"] != "":
                logging.info(f"Skipping row {index} - URL already exists")
                continue

            file_name = sanitize_filename(prompt[:50])
            file_path_local = f"{save_path}/{file_name}.png"

            # Generate image if missing
            if not os.path.exists(file_path_local):
                logging.info(f"Processing prompt {index + 1}: {prompt}")
                generate_image(prompt, save_path)
                new_images_generated = True

            # Upload to Cloudinary only if URL is missing
            if os.path.exists(file_path_local):
                image_url = upload_image(file_path_local)
                if image_url:
                    df.at[index, "Image URL"] = image_url
                    logging.info(f"Image uploaded to Cloudinary. URL: {image_url}")
                    new_images_generated = True

        if new_images_generated:
            df.to_excel(file_path, index=False)
            logging.info("Image URLs updated in Excel.")
        else:
            logging.info("No new images to process.")

    except Exception as e:
        logging.error(f"Error generating images: {e}")
                
def format_excel_columns():
    from openpyxl import load_workbook
    
    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    
    # Format all cells in Media ID column as text
    for row in ws.iter_rows(min_col=4, max_col=4):  # Column D
        for cell in row:
            cell.number_format = '@'
            if cell.value and isinstance(cell.value, float):
                # Convert scientific notation back to full string
                cell.value = f"{int(cell.value)}"
                
    wb.save(EXCEL_FILE_PATH)
    
def process_next_post():
    """
    Processes the FIRST unpublished post in the Excel file.
    Returns True if a post was processed, False if none left.
    """
    try:
        logging.info("Starting post processing...")
        print("\n🔍 Checking for unpublished posts...")
        
        # Read Excel with explicit engine for better error handling
        df = pd.read_excel(EXCEL_FILE_PATH, engine='openpyxl')
        
        # Initialize required columns if missing with proper data types
        required_columns = ['Published', 'Media ID', 'Image URL', 'Generated Captions', 'Publish Date']
        columns_added = False
        
        for col in required_columns:
            if col not in df.columns:
                if col == 'Media ID':
                    df[col] = pd.Series(dtype='string')  # Explicit string type
                elif col == 'Publish Date':
                    df[col] = pd.Series(dtype='datetime64[ns]')
                else:
                    df[col] = "No" if col == 'Published' else ""
                columns_added = True
                logging.warning(f"Added missing column: {col}")

        # Save column changes immediately
        if columns_added:
            df.to_excel(EXCEL_FILE_PATH, index=False, engine='openpyxl')
            logging.info("💾 Saved new column structure to Excel")

        # Process Published column
        if 'Published' in df.columns:
            df['Published'] = df['Published'].fillna("No").replace("", "No")
            df['Media ID'] = df['Media ID'].astype('string')  # Ensure string type
            
        # Find first unpublished post
        unpublished = df[df["Published"] == "No"].head(1)
        
        if unpublished.empty:
            msg = "✅ All posts already published. No action needed."
            logging.info(msg)
            print(msg)
            return False
        
        index = unpublished.index[0]
        row = unpublished.iloc[0]
        print(f"📮 Found unpublished post at row {index+2}")

        # Validate required data
        validation_errors = []
        if pd.isna(row["Image URL"]) or row["Image URL"] == "":
            validation_errors.append("Image URL")
        if pd.isna(row["Generated Captions"]) or row["Generated Captions"] == "":
            validation_errors.append("Generated Captions")
            
        if validation_errors:
            error_msg = f"❌ Missing data in row {index+2}: {', '.join(validation_errors)}"
            logging.error(error_msg)
            print(error_msg)
            return False

        # Publish via Instagram Graph API
        print("🔄 Attempting to publish post...")
        media_id = publish_single_post(row["Image URL"], row["Generated Captions"])
        
        if media_id:
            # Convert media_id to string and add Excel-safe prefix
            media_id_str = f"'{str(media_id).strip()}"  # Apostrophe prefix for Excel string formatting
            
            success_msg = f"📬 Successfully published post! Media ID: {media_id_str}"
            logging.info(success_msg)
            print(success_msg)
            
            
            # Update DataFrame with proper typing
            df['Media ID'] = df['Media ID'].astype('string')
            
            # FIX: Create proper datetime object
            publish_date = pd.to_datetime(datetime.now())
            
            df.at[index, "Published"] = "Yes"
            df.at[index, "Media ID"] = media_id_str
            df.at[index, "Publish Date"] = publish_date 
            
            # Save with explicit string formatting
            df.to_excel(
                EXCEL_FILE_PATH, 
                index=False, 
                engine='openpyxl'
            )
            
            # Apply Excel column formatting
            format_excel_columns()
            
            logging.info(f"💾 Updated Excel file for row {index+2}")
            return True
            
        logging.warning("⚠️ Post publication failed")
        return False

    except Exception as e:
        error_msg = f"🔥 Critical error in process_next_post: {str(e)}"
        logging.error(error_msg, exc_info=True)
        print(error_msg)
        return False
                
# Main function to automate the process
def automate_content_generation():
    logging.info("Starting content generation process...")
    
    # Step 1: Generate captions
    caption_generator = CaptionGenerator(OPENAI_API_KEY)
    caption_result = caption_generator.generate_captions_from_excel(EXCEL_FILE_PATH, COMPANY_NAME)
    logging.info(caption_result)

    # Step 2: Generate images
    generate_images_from_excel(EXCEL_FILE_PATH, IMAGE_SAVE_PATH)
    logging.info("Content generation completed. Ready for publishing.")

# Scheduled post publishing
def publish_scheduled_posts():
    """
    Scheduled task to publish posts from Excel.
    """
    logging.info("Running scheduled post publishing...")
    processed = process_next_post()
    if processed:
        logging.info("Published 1 new post.")
    else:
        logging.info("No new posts to publish.")
        
# Schedule post publishing every 3 minutes
schedule.every(1).minutes.do(publish_scheduled_posts)
schedule.every(3).minutes.do(lambda: collect_analytics(days=30))

# Run the script directly
if __name__ == "__main__":
    logging.info("Script started.")
    automate_content_generation()
    # Run the scheduler
    while True:
        try:
            schedule.run_pending()
            time.sleep(1)
        except Exception as e:
            logging.error(f"Error in scheduled task: {str(e)}")
            time.sleep(60)